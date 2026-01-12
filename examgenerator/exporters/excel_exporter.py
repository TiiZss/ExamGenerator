"""
Excel exporter for exam answers.
"""

import os
from datetime import datetime
from typing import List, Dict, Any, cast


def create_answers_excel(all_exam_data: List[Dict[str, Any]], exam_prefix: str, output_dir: str, minutes_per_question: float = 1.0) -> None:
    """Create a single Excel file with all exam answers (transposed layout)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Error: Para exportar a Excel necesitas instalar openpyxl:")
        print("uv add openpyxl")
        return

    # Import time calculation function
    from ..core.time_calculator import calculate_exam_time

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)  # Type hint para evitar errores de Pylance
    ws.title = f"Respuestas {exam_prefix}"
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    exam_font = Font(bold=True, size=11)
    exam_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    answer_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Find maximum number of questions across all exams
    max_questions = max(len(exam_data['answers']) for exam_data in all_exam_data)
    
    # Set up headers (TRANSPOSED: Examen in rows, Questions in columns)
    headers = ["Examen"]
    for i in range(1, max_questions + 1):
        headers.append(f"P{i}")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write exam data (TRANSPOSED: each row is an exam)
    for exam_idx, exam_data in enumerate(all_exam_data):
        row = exam_idx + 2
        
        # Exam name
        cell = ws.cell(row=row, column=1, value=f"Examen {exam_data['exam_number']}")
        cell.font = exam_font
        cell.fill = exam_fill
        cell.alignment = answer_alignment
        cell.border = thin_border
        
        # Answers for each question
        for question_idx in range(max_questions):
            col = question_idx + 2
            if question_idx < len(exam_data['answers']):
                answer = exam_data['answers'][question_idx]
                cell = ws.cell(row=row, column=col, value=answer)
            else:
                cell = ws.cell(row=row, column=col, value="-")
            
            cell.alignment = answer_alignment
            cell.border = thin_border
    
    # Add exam info at the bottom
    info_start_row = len(all_exam_data) + 4
    
    # Exam details header
    cell = ws.cell(row=info_start_row, column=1, value="Información de Exámenes")
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    
    # Merge cells for the header
    ws.merge_cells(start_row=info_start_row, start_column=1, 
                  end_row=info_start_row, end_column=len(headers))
    
    # Exam details
    details = [
        ("Nombre del examen:", exam_prefix),
        ("Fecha de generación:", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Número total de exámenes:", str(len(all_exam_data))),
        ("Preguntas por examen:", str(max_questions)),
        ("Tiempo estimado:", calculate_exam_time(max_questions, minutes_per_question))
    ]
    
    for i, (label, value) in enumerate(details):
        row = info_start_row + i + 1
        
        # Label
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = exam_font
        cell.border = thin_border
        
        # Value
        cell = ws.cell(row=row, column=2, value=value)
        cell.border = thin_border
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        if col == 1:
            ws.column_dimensions[column_letter].width = 15
        else:
            ws.column_dimensions[column_letter].width = 8
    
    # Save the workbook
    excel_filename = os.path.join(output_dir, f"respuestas_{exam_prefix}_completas.xlsx")
    wb.save(excel_filename)
    print(f"Archivo Excel creado: {excel_filename}")
